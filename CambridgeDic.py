import requests, sys
from lxml import etree
from openpyxl import load_workbook
import os

def Dic(word):
    r = requests.get('https://dictionary.cambridge.org/dictionary/english-chinese-traditional/' + word, headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36'})
    
    # parse html
    html = r.text
    page = etree.HTML(html)
    # eng def
    endefs = page.xpath('//*[@id="page-content"]/div[2]/div[4]/div/div/div[1]/div[3]/div/div[2]/div[1]/div[2]/div')
    endefstr = endefs[0].xpath('string(.)').strip()
    # chi def
    # xpath will return a list
    chdef = page.xpath('//*[@id="page-content"]/div[2]/div[4]/div/div/div[1]/div[3]/div/div[2]/div[1]/div[3]/span')
    chdefstr = chdef[0].xpath('string(.)').strip()
    print('Definition:\n',endefstr, '\n',chdefstr)

    # eng sentence
    enstc = page.xpath('//*[@id="page-content"]/div[2]/div[4]/div/div/div[1]/div[3]/div/div[2]/div[1]/div[3]/div[1]/span[1]')
    enstcstr = enstc[0].xpath('string(.)').strip()
    # chi sentence
    chstc = page.xpath('//*[@id="page-content"]/div[2]/div[4]/div/div/div[1]/div[3]/div/div[2]/div[1]/div[3]/div[1]/span[2]')
    chstcstr = chstc[0].xpath('string(.)').strip()

    print('Sentence:\n', enstcstr,'\n', chstcstr)
    
    #store the word
    isstore = input('Would you like to store? y or n: ')
    
    if isstore=='y':
        StoreExcel(word, chdefstr, enstcstr)
    

def StoreExcel(word, chdefstr, enstcstr):

    wb = load_workbook('voc_cambridge.xlsx')
    sheet = wb.get_sheet_by_name('Sheet')
    currentCount = sheet.max_row
    # get position
    wordPos = 'A' + str(currentCount+1)
    defPos = 'B' + str(currentCount+1)
    senPos = 'C' + str(currentCount+1)
    # write new word
    sheet[wordPos] = word
    sheet[defPos] = chdefstr
    sheet[senPos] = enstcstr
    # numbers of words add 1
    sheet['E1'] = currentCount + 1
    
    # save file
    wb.save('voc_cambridge.xlsx')
    

if len(sys.argv) > 1:
    word = sys.argv[1]
    Dic(word)
else:
    print('No word!')

